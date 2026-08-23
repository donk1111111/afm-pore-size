#!/usr/bin/env python3
"""Measure complete circular pores in AFM-rendered TIFF images.

The script is intentionally dependency-light for desktop use: Pillow, NumPy,
OpenCV, and scikit-image are enough for the analysis path.
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from skimage import measure, morphology
from skimage.feature import blob_log


warnings.filterwarnings("ignore", category=FutureWarning)

IMAGE_EXTENSIONS = (".tif", ".tiff", ".png", ".jpg", ".jpeg", ".bmp")


@dataclass
class ScaleInfo:
    unit_per_px: float | None
    unit: str | None
    source: str
    scale_bar_px: float | None = None
    scale_bar_bbox: tuple[int, int, int, int] | None = None


@dataclass
class Crop:
    x: int
    y: int
    w: int
    h: int

    @property
    def bounds(self) -> tuple[int, int, int, int]:
        return self.x, self.y, self.x + self.w, self.y + self.h


def parse_box(text: str) -> tuple[int, int, int, int]:
    parts = [int(p.strip()) for p in text.split(",")]
    if len(parts) != 4:
        raise argparse.ArgumentTypeError("Expected x,y,w,h")
    if parts[2] <= 0 or parts[3] <= 0:
        raise argparse.ArgumentTypeError("Width and height must be positive")
    return tuple(parts)  # type: ignore[return-value]


def normalize_unit(unit: str | None) -> str | None:
    if unit is None:
        return None
    value = unit.strip().lower()
    aliases = {
        "micron": "um",
        "microns": "um",
        "micro": "um",
        "micrometer": "um",
        "micrometers": "um",
        "mum": "um",
        "nanometer": "nm",
        "nanometers": "nm",
        "millimeter": "mm",
        "millimeters": "mm",
    }
    return aliases.get(value, value)


def collect_images(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if not input_path.is_dir():
        raise FileNotFoundError(f"Input path does not exist: {input_path}")
    images: list[Path] = []
    for path in sorted(input_path.rglob("*")):
        suffix = path.suffix.lower()
        name = path.name.lower()
        if suffix in IMAGE_EXTENSIONS or name.endswith(".spm.tif"):
            images.append(path)
    return images


def load_rgb(path: Path) -> np.ndarray:
    with Image.open(path) as im:
        return np.array(im.convert("RGB"))


def gray_from_rgb(rgb: np.ndarray) -> np.ndarray:
    return cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY).astype(np.uint8)


def auto_content_crop(rgb: np.ndarray, pad: int = 2) -> Crop:
    gray = gray_from_rgb(rgb)
    # Rendered AFM images often have white margins. Keep annotations inside the
    # crop; shape filters will reject text and scale bars later.
    non_white = gray < 248
    rows = np.where(non_white.any(axis=1))[0]
    cols = np.where(non_white.any(axis=0))[0]
    h, w = gray.shape
    if rows.size == 0 or cols.size == 0:
        return Crop(0, 0, w, h)
    x0 = max(int(cols.min()) - pad, 0)
    y0 = max(int(rows.min()) - pad, 0)
    x1 = min(int(cols.max()) + pad + 1, w)
    y1 = min(int(rows.max()) + pad + 1, h)
    return Crop(x0, y0, x1 - x0, y1 - y0)


def crop_array(array: np.ndarray, crop: Crop) -> np.ndarray:
    return array[crop.y : crop.y + crop.h, crop.x : crop.x + crop.w]


def circle_crosses_image_edge(x: float, y: float, radius: float, width: int, height: int) -> bool:
    safety_px = max(2.0, 0.05 * radius)
    return (
        x - radius < safety_px
        or x + radius > width - safety_px
        or y - radius < safety_px
        or y + radius > height - safety_px
    )


def circle_edge_gap(x: float, y: float, radius: float, width: int, height: int) -> float:
    return min(
        x - radius,
        y - radius,
        width - (x + radius),
        height - (y + radius),
    )


def scale_bar_candidates(rgb: np.ndarray, roi: tuple[int, int, int, int] | None) -> list[tuple[float, tuple[int, int, int, int], str]]:
    gray = gray_from_rgb(rgb)
    h, w = gray.shape
    if roi:
        x0, y0, rw, rh = roi
    else:
        x0, y0, rw, rh = 0, int(h * 0.62), w, h - int(h * 0.62)
    x0 = max(0, min(x0, w - 1))
    y0 = max(0, min(y0, h - 1))
    x1 = max(x0 + 1, min(x0 + rw, w))
    y1 = max(y0 + 1, min(y0 + rh, h))
    roi_gray = gray[y0:y1, x0:x1]

    masks = [
        ("dark", roi_gray < 80),
        ("bright", roi_gray > 235),
    ]
    candidates: list[tuple[float, tuple[int, int, int, int], str]] = []
    for polarity, mask in masks:
        labels = measure.label(mask.astype(bool), connectivity=2)
        props = measure.regionprops(labels)
        for prop in props:
            minr, minc, maxr, maxc = prop.bbox
            bw = maxc - minc
            bh = maxr - minr
            area = prop.area
            if bw < 20 or bh < 2:
                continue
            aspect = bw / max(bh, 1)
            fill = area / max(bw * bh, 1)
            if aspect < 4.0 or bh > 25 or fill < 0.35:
                continue
            abs_x = x0 + minc
            abs_y = y0 + minr
            bottomness = (abs_y + bh / 2) / h
            rightness = (abs_x + bw / 2) / w
            # Prefer long, thin, solid bars near the bottom/right, but do not
            # require that location because some AFM exports place bars left.
            score = bw * 2.0 + aspect * 8.0 + fill * 40.0 + bottomness * 30.0 + rightness * 8.0 - bh
            candidates.append((score, (abs_x, abs_y, bw, bh), polarity))
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates


def detect_scale_info(
    rgb: np.ndarray,
    scale_bar_value: float | None,
    scale_bar_unit: str | None,
    scale_bar_roi: tuple[int, int, int, int] | None,
    unit_per_px: float | None,
    unit: str | None,
) -> ScaleInfo:
    unit = normalize_unit(unit)
    scale_bar_unit = normalize_unit(scale_bar_unit)
    if unit_per_px is not None:
        if not unit:
            raise ValueError("--unit is required when --unit-per-px is used")
        candidates = scale_bar_candidates(rgb, scale_bar_roi)
        bbox = candidates[0][1] if candidates else None
        px = float(bbox[2]) if bbox else None
        return ScaleInfo(unit_per_px, unit, "unit_per_px", px, bbox)

    candidates = scale_bar_candidates(rgb, scale_bar_roi)
    bbox = candidates[0][1] if candidates else None
    px = float(bbox[2]) if bbox else None
    if scale_bar_value is not None:
        if not scale_bar_unit:
            raise ValueError("--scale-bar-unit is required with --scale-bar-value")
        if not px or px <= 0:
            raise ValueError("Could not detect a scale bar; use --scale-bar-roi or --unit-per-px")
        return ScaleInfo(scale_bar_value / px, scale_bar_unit, "scale_bar", px, bbox)
    return ScaleInfo(None, None, "pixels", px, bbox)


def remove_scale_bar(mask: np.ndarray, crop: Crop, scale_bbox: tuple[int, int, int, int] | None) -> None:
    if scale_bbox is None:
        return
    x, y, w, h = scale_bbox
    x0 = max(0, x - crop.x - 8)
    y0 = max(0, y - crop.y - 8)
    x1 = min(mask.shape[1], x - crop.x + w + 8)
    y1 = min(mask.shape[0], y - crop.y + h + 8)
    if x0 < x1 and y0 < y1:
        mask[y0:y1, x0:x1] = False


def threshold_pores(gray: np.ndarray, polarity: str) -> np.ndarray:
    smooth = cv2.GaussianBlur(gray, (0, 0), 9)
    if polarity == "dark":
        response = smooth.astype(np.int16) - gray.astype(np.int16)
    elif polarity == "bright":
        response = gray.astype(np.int16) - smooth.astype(np.int16)
    else:
        dark = threshold_pores(gray, "dark")
        bright = threshold_pores(gray, "bright")
        # AFM pores are usually dark depressions. Keep bright candidates too,
        # but the later shape filters decide what survives.
        return np.logical_or(dark, bright)

    response = np.clip(response, 0, 255).astype(np.uint8)
    if int(response.max()) <= 0:
        return np.zeros_like(gray, dtype=bool)
    otsu, _ = cv2.threshold(response, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    percentile = float(np.percentile(response, 93))
    thresh = max(float(otsu), percentile, 8.0)
    mask = response >= thresh
    mask = morphology.remove_small_objects(mask, min_size=20)
    mask = morphology.binary_opening(mask, morphology.disk(1))
    mask = morphology.binary_closing(mask, morphology.disk(2))
    mask = morphology.remove_small_holes(mask, area_threshold=10000)
    return mask.astype(bool)


def feret_diameter(points_xy: np.ndarray) -> float:
    if len(points_xy) < 2:
        return 0.0
    hull = cv2.convexHull(points_xy.astype(np.float32))
    pts = hull.reshape(-1, 2)
    if len(pts) < 2:
        return 0.0
    # Pore contours are small enough that pairwise hull distances are fine and
    # avoid a more fragile rotating-calipers implementation.
    diff = pts[:, None, :] - pts[None, :, :]
    dist2 = np.sum(diff * diff, axis=2)
    return float(np.sqrt(dist2.max()))


def ellipse_metrics(coords_rc: np.ndarray) -> tuple[float, float, float, float]:
    points_xy = np.column_stack([coords_rc[:, 1], coords_rc[:, 0]]).astype(np.float32)
    if len(points_xy) < 5:
        return 1.0, 1.0, 0.0, 0.0
    try:
        (_, _), (axis_a, axis_b), angle = cv2.fitEllipse(points_xy)
    except cv2.error:
        return 1.0, 1.0, 0.0, 0.0
    major = max(float(axis_a), float(axis_b))
    minor = min(float(axis_a), float(axis_b))
    if major <= 0 or minor <= 0:
        return 1.0, 1.0, 0.0, float(angle)
    ar = major / minor
    roundness = minor / major
    return ar, roundness, minor, float(angle)


def contour_for_region(label_img: np.ndarray, label: int) -> np.ndarray:
    component = (label_img == label).astype(np.uint8)
    contours, _ = cv2.findContours(component, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not contours:
        return np.empty((0, 1, 2), dtype=np.int32)
    return max(contours, key=cv2.contourArea)


def data_crop_for_pores(content_crop: Crop, scale_info: ScaleInfo) -> Crop:
    if scale_info.scale_bar_bbox:
        _, scale_y, _, _ = scale_info.scale_bar_bbox
        scan_size = max(1, scale_y - content_crop.y - 8)
        scan_size = min(scan_size, content_crop.w, content_crop.h)
        return Crop(content_crop.x, content_crop.y, scan_size, scan_size)
    scan_size = min(content_crop.w, content_crop.h)
    return Crop(content_crop.x, content_crop.y, scan_size, scan_size)


def scaled_fields(
    label: int,
    area_px: float,
    perim_px: float,
    circularity: float,
    feret_px: float,
    center_x_px: float,
    center_y_px: float,
    angle: float,
    min_feret_px: float,
    ar: float,
    roundness: float,
    solidity: float,
    bbox: tuple[int, int, int, int],
    scale_info: ScaleInfo,
) -> dict[str, float | int | str]:
    unit_per_px = scale_info.unit_per_px
    unit_name = scale_info.unit or "px"

    def scaled(value: float) -> float:
        return value * unit_per_px if unit_per_px is not None else value

    row: dict[str, float | int | str] = {
        "Label": label,
        "Area": area_px * unit_per_px * unit_per_px if unit_per_px is not None else area_px,
        "Perim.": scaled(perim_px),
        "Circ.": circularity,
        "Feret": scaled(feret_px),
        "FeretX": scaled(center_x_px),
        "FeretY": scaled(center_y_px),
        "FeretAngle": angle,
        "MinFeret": scaled(min_feret_px),
        "AR": ar,
        "Round": roundness,
        "Solidity": solidity,
        "unit": unit_name,
        "Area_px": area_px,
        "Perim_px": perim_px,
        "Feret_px": feret_px,
        "Diameter_px": 2.0 * math.sqrt(area_px / math.pi),
        "MinFeret_px": min_feret_px,
        "CenterX_px": center_x_px,
        "CenterY_px": center_y_px,
        "BBoxX_px": bbox[0],
        "BBoxY_px": bbox[1],
        "BBoxW_px": bbox[2],
        "BBoxH_px": bbox[3],
    }
    if unit_per_px is not None and unit_name:
        row[f"Diameter_{unit_name}"] = scaled(float(row["Diameter_px"]))
        row[f"Feret_{unit_name}"] = scaled(feret_px)
        row[f"Area_{unit_name}2"] = area_px * unit_per_px * unit_per_px
    return row


def detect_blob_pores(
    rgb: np.ndarray,
    data_crop: Crop,
    scale_info: ScaleInfo,
    args: argparse.Namespace,
) -> tuple[list[dict[str, float | int | str]], np.ndarray, list[dict[str, float | int | str]]]:
    data = crop_array(rgb, data_crop)
    gray = gray_from_rgb(data).astype(float)
    if args.polarity == "bright":
        response = gray
    else:
        response = 255.0 - gray
    response = response - response.min()
    max_value = response.max()
    if max_value > 0:
        response = response / max_value

    min_radius = max(args.min_diameter_px / 2.0, 3.0)
    max_radius = (args.max_diameter_px / 2.0) if args.max_diameter_px else min(data_crop.w, data_crop.h) / 5.0
    blobs = blob_log(
        response,
        min_sigma=max(min_radius / math.sqrt(2.0), 1.0),
        max_sigma=max(max_radius / math.sqrt(2.0), 2.0),
        num_sigma=12,
        threshold=args.blob_threshold,
        overlap=args.blob_overlap,
    )

    rows: list[dict[str, float | int | str]] = []
    candidate_rows: list[dict[str, float | int | str]] = []
    mask = np.zeros(response.shape, dtype=bool)
    for blob in blobs:
        y, x, sigma = [float(v) for v in blob]
        radius = sigma * math.sqrt(2.0)
        feret_px = 2.0 * radius
        abs_x = data_crop.x + x
        abs_y = data_crop.y + y

        reject_reason = ""
        if feret_px < args.min_diameter_px:
            reject_reason = "min_diameter"
        elif args.max_diameter_px and feret_px > args.max_diameter_px:
            reject_reason = "max_diameter"
        elif circle_crosses_image_edge(x, y, radius, data_crop.w, data_crop.h):
            reject_reason = "edge_margin"
        elif circle_edge_gap(x, y, radius, data_crop.w, data_crop.h) < max(4.0, 0.15 * radius):
            reject_reason = "edge_fragment"
        score_y = max(0, min(int(round(y)), response.shape[0] - 1))
        score_x = max(0, min(int(round(x)), response.shape[1] - 1))
        if scale_info.unit_per_px is not None:
            center_x = abs_x * scale_info.unit_per_px
            center_y = abs_y * scale_info.unit_per_px
            diameter_um = feret_px * scale_info.unit_per_px
        else:
            center_x = abs_x
            center_y = abs_y
            diameter_um = ""
        candidate_rows.append(
            {
                "candidate_id": len(candidate_rows) + 1,
                "center_x": center_x,
                "center_y": center_y,
                "radius_px": radius,
                "diameter_um": diameter_um,
                "sigma": sigma,
                "score": float(response[score_y, score_x]),
                "filtered": bool(reject_reason),
                "reject_reason": reject_reason,
            }
        )
        if reject_reason:
            continue

        bbox = (
            int(round(abs_x - radius)),
            int(round(abs_y - radius)),
            int(round(2 * radius)),
            int(round(2 * radius)),
        )
        area_px = math.pi * radius * radius
        perim_px = 2.0 * math.pi * radius
        row = scaled_fields(
            len(rows) + 1,
            area_px,
            perim_px,
            1.0,
            feret_px,
            abs_x,
            abs_y,
            0.0,
            feret_px,
            1.0,
            1.0,
            1.0,
            bbox,
            scale_info,
        )
        rows.append(row)
        cv2.circle(mask, (int(round(x)), int(round(y))), int(round(radius)), True, -1)
    return rows, mask, candidate_rows


def detect_small_blob_candidates(
    rgb: np.ndarray,
    data_crop: Crop,
    scale_info: ScaleInfo,
    args: argparse.Namespace,
) -> list[dict[str, float | int | str]]:
    data = crop_array(rgb, data_crop)
    gray = gray_from_rgb(data).astype(float)
    if args.polarity == "bright":
        response = gray
    else:
        response = 255.0 - gray
    response = response - response.min()
    max_value = response.max()
    if max_value > 0:
        response = response / max_value

    min_radius = max(args.small_blob_min_diameter_px / 2.0, 1.0)
    max_radius = max(args.small_blob_max_diameter_px / 2.0, min_radius)
    blobs = blob_log(
        response,
        min_sigma=max(min_radius / math.sqrt(2.0), 1.0),
        max_sigma=max(max_radius / math.sqrt(2.0), 1.0),
        num_sigma=12,
        threshold=args.blob_threshold,
        overlap=args.blob_overlap,
    )

    candidate_rows: list[dict[str, float | int | str]] = []
    for blob in blobs:
        y, x, sigma = [float(v) for v in blob]
        radius = sigma * math.sqrt(2.0)
        feret_px = 2.0 * radius
        abs_x = data_crop.x + x
        abs_y = data_crop.y + y
        score_y = max(0, min(int(round(y)), response.shape[0] - 1))
        score_x = max(0, min(int(round(x)), response.shape[1] - 1))
        if scale_info.unit_per_px is not None:
            center_x = abs_x * scale_info.unit_per_px
            center_y = abs_y * scale_info.unit_per_px
            diameter_um = feret_px * scale_info.unit_per_px
        else:
            center_x = abs_x
            center_y = abs_y
            diameter_um = ""
        candidate_rows.append(
            {
                "candidate_id": len(candidate_rows) + 1,
                "center_x": center_x,
                "center_y": center_y,
                "radius_px": radius,
                "diameter_um": diameter_um,
                "score": float(response[score_y, score_x]),
            }
        )
    return candidate_rows


def candidate_center_px(candidate: dict[str, float | int | str], scale_info: ScaleInfo) -> tuple[float, float]:
    center_x = float(candidate["center_x"])
    center_y = float(candidate["center_y"])
    if scale_info.unit_per_px:
        return center_x / scale_info.unit_per_px, center_y / scale_info.unit_per_px
    return center_x, center_y


def local_dark_contrast(gray: np.ndarray, x: float, y: float, radius: float) -> float:
    yy, xx = np.ogrid[: gray.shape[0], : gray.shape[1]]
    distances = (xx - x) ** 2 + (yy - y) ** 2
    center_mask = distances <= (0.35 * radius) ** 2
    ring_mask = np.logical_and(distances >= (0.90 * radius) ** 2, distances <= (1.50 * radius) ** 2)
    if not center_mask.any() or not ring_mask.any():
        return 0.0
    return float(np.mean(gray[ring_mask]) - np.mean(gray[center_mask]))


def radial_circularity_score(gray: np.ndarray, x: float, y: float, radius: float) -> float:
    sampled_radii: list[float] = []
    for index in range(24):
        theta = 2.0 * math.pi * index / 24.0
        radial_points = []
        for sample_radius in np.linspace(max(1.0, 0.35 * radius), 1.80 * radius, 24):
            sample_x = int(round(x + math.cos(theta) * sample_radius))
            sample_y = int(round(y + math.sin(theta) * sample_radius))
            if 0 <= sample_x < gray.shape[1] and 0 <= sample_y < gray.shape[0]:
                radial_points.append((sample_radius, float(gray[sample_y, sample_x])))
        if len(radial_points) < 6:
            continue
        center_level = float(np.mean([value for _, value in radial_points[:3]]))
        outer_level = float(np.mean([value for _, value in radial_points[-5:]]))
        threshold = center_level + max(3.0, 0.5 * (outer_level - center_level))
        for sample_radius, value in radial_points:
            if value >= threshold:
                sampled_radii.append(sample_radius)
                break
    if len(sampled_radii) < 8:
        return 0.0
    radius_mean = float(np.mean(sampled_radii))
    radius_std = float(np.std(sampled_radii))
    return max(0.0, 1.0 - radius_std / max(radius_mean, 1.0))


def nearest_main_pore_distance_px(
    x: float,
    y: float,
    radius: float,
    main_rows: list[dict[str, float | int | str]],
) -> tuple[float, float]:
    nearest_distance = math.inf
    nearest_duplicate_limit = 0.0
    for row in main_rows:
        main_x = float(row["CenterX_px"])
        main_y = float(row["CenterY_px"])
        main_radius = float(row["Diameter_px"]) / 2.0
        distance = math.hypot(x - main_x, y - main_y)
        duplicate_limit = max(8.0, main_radius + radius)
        if distance < nearest_distance:
            nearest_distance = distance
            nearest_duplicate_limit = duplicate_limit
    return nearest_distance, nearest_duplicate_limit


def validate_small_blob_candidates(
    rgb: np.ndarray,
    data_crop: Crop,
    scale_info: ScaleInfo,
    main_rows: list[dict[str, float | int | str]],
    small_candidates: list[dict[str, float | int | str]],
    args: argparse.Namespace,
) -> tuple[list[dict[str, float | int | str]], list[dict[str, float | int | str]]]:
    gray = gray_from_rgb(rgb).astype(float)
    validated_rows: list[dict[str, float | int | str]] = []
    accepted_rows: list[dict[str, float | int | str]] = []
    for candidate in small_candidates:
        abs_x, abs_y = candidate_center_px(candidate, scale_info)
        radius = float(candidate["radius_px"])
        crop_x = abs_x - data_crop.x
        crop_y = abs_y - data_crop.y
        edge_gap = circle_edge_gap(crop_x, crop_y, radius, data_crop.w, data_crop.h)
        contrast = local_dark_contrast(gray, abs_x, abs_y, radius)
        circularity_score = radial_circularity_score(gray, abs_x, abs_y, radius)
        main_distance, duplicate_limit = nearest_main_pore_distance_px(abs_x, abs_y, radius, main_rows)

        reject_reason = ""
        if main_distance <= duplicate_limit:
            reject_reason = "duplicate"
        elif circle_crosses_image_edge(crop_x, crop_y, radius, data_crop.w, data_crop.h):
            reject_reason = "edge_margin"
        elif edge_gap < max(4.0, 0.15 * radius):
            reject_reason = "edge_fragment"
        elif float(candidate["score"]) < args.small_blob_min_score:
            reject_reason = "low_score"
        elif contrast < args.small_blob_min_contrast:
            reject_reason = "low_contrast"
        elif circularity_score < args.small_blob_min_circularity_score:
            reject_reason = "low_circularity"

        accepted = not reject_reason
        validated_rows.append(
            {
                "candidate_id": candidate["candidate_id"],
                "center_x": candidate["center_x"],
                "center_y": candidate["center_y"],
                "diameter_um": candidate["diameter_um"],
                "contrast": contrast,
                "circularity_score": circularity_score,
                "accepted": accepted,
                "reject_reason": reject_reason,
            }
        )
        if not accepted:
            continue

        feret_px = 2.0 * radius
        bbox = (
            int(round(abs_x - radius)),
            int(round(abs_y - radius)),
            int(round(feret_px)),
            int(round(feret_px)),
        )
        accepted_rows.append(
            scaled_fields(
                len(main_rows) + len(accepted_rows) + 1,
                math.pi * radius * radius,
                2.0 * math.pi * radius,
                1.0,
                feret_px,
                abs_x,
                abs_y,
                0.0,
                feret_px,
                1.0,
                1.0,
                1.0,
                bbox,
                scale_info,
            )
        )
    return validated_rows, accepted_rows


def analyze_image(rgb: np.ndarray, args: argparse.Namespace) -> tuple[list[dict[str, float | int | str]], Crop, ScaleInfo, np.ndarray, list[dict[str, float | int | str]]]:
    crop = Crop(*args.crop) if args.crop else auto_content_crop(rgb)
    cropped = crop_array(rgb, crop)
    gray = gray_from_rgb(cropped)

    scale_info = detect_scale_info(
        rgb,
        args.scale_bar_value,
        args.scale_bar_unit,
        args.scale_bar_roi,
        args.unit_per_px,
        args.unit,
    )

    if args.method == "blob":
        blob_crop = data_crop_for_pores(crop, scale_info)
        blob_rows, blob_mask, candidate_rows = detect_blob_pores(rgb, blob_crop, scale_info, args)
        return blob_rows, blob_crop, scale_info, blob_mask, candidate_rows

    mask = threshold_pores(gray, args.polarity)
    remove_scale_bar(mask, crop, scale_info.scale_bar_bbox)
    mask = morphology.remove_small_objects(mask, min_size=max(5, args.min_area_px))
    label_img = measure.label(mask, connectivity=2)
    props = measure.regionprops(label_img)

    rows: list[dict[str, float | int | str]] = []
    for prop in props:
        minr, minc, maxr, maxc = prop.bbox
        if minr <= args.edge_margin_px or minc <= args.edge_margin_px:
            continue
        if maxr >= crop.h - args.edge_margin_px or maxc >= crop.w - args.edge_margin_px:
            continue

        area_px = float(prop.area)
        contour = contour_for_region(label_img, prop.label)
        if contour.size == 0:
            continue
        perim_px = float(cv2.arcLength(contour, True))
        if perim_px <= 0:
            continue
        circularity = float(4.0 * math.pi * area_px / (perim_px * perim_px))
        if circularity < args.min_circularity:
            continue

        coords = prop.coords
        ar, roundness, min_feret_px, angle = ellipse_metrics(coords)
        if ar > args.max_ar or roundness < args.min_round:
            continue

        solidity = float(prop.solidity) if prop.solidity else 0.0
        if solidity < args.min_solidity:
            continue

        points_xy = np.column_stack([coords[:, 1], coords[:, 0]])
        feret_px = max(feret_diameter(points_xy), 2.0 * math.sqrt(area_px / math.pi))
        if min_feret_px <= 0:
            min_feret_px = min(float(maxr - minr), float(maxc - minc))
        diameter_px = 2.0 * math.sqrt(area_px / math.pi)
        if feret_px < args.min_diameter_px:
            continue
        if args.max_diameter_px and feret_px > args.max_diameter_px:
            continue

        centroid_y, centroid_x = prop.centroid
        abs_x = float(crop.x + centroid_x)
        abs_y = float(crop.y + centroid_y)
        unit_per_px = scale_info.unit_per_px
        unit_name = scale_info.unit or "px"

        def scaled(value: float) -> float:
            return value * unit_per_px if unit_per_px is not None else value

        row: dict[str, float | int | str] = {
            "Label": len(rows) + 1,
            "Area": area_px * unit_per_px * unit_per_px if unit_per_px is not None else area_px,
            "Perim.": scaled(perim_px),
            "Circ.": circularity,
            "Feret": scaled(feret_px),
            "FeretX": scaled(abs_x),
            "FeretY": scaled(abs_y),
            "FeretAngle": angle,
            "MinFeret": scaled(min_feret_px),
            "AR": ar,
            "Round": roundness,
            "Solidity": solidity,
            "unit": unit_name,
            "Area_px": area_px,
            "Perim_px": perim_px,
            "Feret_px": feret_px,
            "Diameter_px": diameter_px,
            "MinFeret_px": min_feret_px,
            "CenterX_px": abs_x,
            "CenterY_px": abs_y,
            "BBoxX_px": crop.x + int(minc),
            "BBoxY_px": crop.y + int(minr),
            "BBoxW_px": int(maxc - minc),
            "BBoxH_px": int(maxr - minr),
        }
        if unit_per_px is not None and unit_name:
            row[f"Diameter_{unit_name}"] = scaled(diameter_px)
            row[f"Feret_{unit_name}"] = scaled(feret_px)
            row[f"Area_{unit_name}2"] = area_px * unit_per_px * unit_per_px
        rows.append(row)

    if args.method == "auto" and not rows:
        blob_crop = data_crop_for_pores(crop, scale_info)
        blob_rows, blob_mask, candidate_rows = detect_blob_pores(rgb, blob_crop, scale_info, args)
        return blob_rows, blob_crop, scale_info, blob_mask, candidate_rows

    return rows, crop, scale_info, mask, []


def annotate_image(
    rgb: np.ndarray,
    rows: list[dict[str, float | int | str]],
    crop: Crop,
    scale_info: ScaleInfo,
    out_path: Path,
) -> None:
    im = Image.fromarray(rgb.copy())
    draw = ImageDraw.Draw(im)
    try:
        font = ImageFont.load_default()
    except Exception:
        font = None

    draw.rectangle([crop.x, crop.y, crop.x + crop.w, crop.y + crop.h], outline=(0, 255, 255), width=2)
    if scale_info.scale_bar_bbox:
        x, y, w, h = scale_info.scale_bar_bbox
        draw.rectangle([x - 3, y - 3, x + w + 3, y + h + 3], outline=(255, 0, 0), width=2)
        label = f"{scale_info.scale_bar_px:.1f}px" if scale_info.scale_bar_px else "scale"
        if scale_info.unit_per_px and scale_info.unit:
            label += f" = {scale_info.unit_per_px:.6g} {scale_info.unit}/px"
        draw.text((x, max(0, y - 14)), label, fill=(255, 0, 0), font=font)

    for row in rows:
        cx = float(row["CenterX_px"])
        cy = float(row["CenterY_px"])
        radius = float(row["Feret_px"]) / 2.0
        draw.ellipse([cx - radius, cy - radius, cx + radius, cy + radius], outline=(0, 255, 0), width=2)
        draw.text((cx + 4, cy + 4), str(row["Label"]), fill=(0, 255, 0), font=font)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    im.save(out_path)


def write_mask(mask: np.ndarray, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    Image.fromarray((mask.astype(np.uint8) * 255)).save(out_path)


def fieldnames(rows: list[dict[str, float | int | str]]) -> list[str]:
    preferred = [
        "Image",
        "Label",
        "Area",
        "Perim.",
        "Circ.",
        "Feret",
        "FeretX",
        "FeretY",
        "FeretAngle",
        "MinFeret",
        "AR",
        "Round",
        "Solidity",
        "unit",
        "scale_source",
        "scale_bar_px",
        "unit_per_px",
        "Area_px",
        "Perim_px",
        "Feret_px",
        "Diameter_px",
        "MinFeret_px",
        "CenterX_px",
        "CenterY_px",
        "BBoxX_px",
        "BBoxY_px",
        "BBoxW_px",
        "BBoxH_px",
    ]
    extra: list[str] = []
    for row in rows:
        for key in row:
            if key not in preferred and key not in extra:
                extra.append(key)
    return [name for name in preferred if any(name in r for r in rows)] + extra


def write_csv(path: Path, rows: list[dict[str, float | int | str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        if not rows:
            handle.write("")
            return
        writer = csv.DictWriter(handle, fieldnames=fieldnames(rows))
        writer.writeheader()
        writer.writerows(rows)


def sample_gray(gray: np.ndarray, x: float, y: float) -> float | None:
    if x < 0 or y < 0 or x >= gray.shape[1] - 1 or y >= gray.shape[0] - 1:
        return None
    x0 = int(math.floor(x))
    y0 = int(math.floor(y))
    dx = x - x0
    dy = y - y0
    top = (1.0 - dx) * gray[y0, x0] + dx * gray[y0, x0 + 1]
    bottom = (1.0 - dx) * gray[y0 + 1, x0] + dx * gray[y0 + 1, x0 + 1]
    return float((1.0 - dy) * top + dy * bottom)


def refine_radius_for_row(gray: np.ndarray, row: dict[str, float | int | str]) -> tuple[float, float, float]:
    center_x = float(row["CenterX_px"])
    center_y = float(row["CenterY_px"])
    old_radius = float(row["Diameter_px"]) / 2.0
    candidates: list[float] = []
    for index in range(72):
        theta = 2.0 * math.pi * index / 72.0
        radii = np.linspace(max(1.0, 0.35 * old_radius), max(2.0, 1.80 * old_radius), 80)
        values: list[float] = []
        valid_radii: list[float] = []
        for radius in radii:
            value = sample_gray(gray, center_x + math.cos(theta) * radius, center_y + math.sin(theta) * radius)
            if value is None:
                continue
            valid_radii.append(float(radius))
            values.append(value)
        if len(values) < 8:
            continue
        smoothed = np.convolve(np.array(values, dtype=float), np.ones(5) / 5.0, mode="same")
        gradients = np.gradient(smoothed)
        search_start = max(1, int(0.15 * len(gradients)))
        search_end = max(search_start + 1, int(0.95 * len(gradients)))
        local_index = int(np.argmax(gradients[search_start:search_end]) + search_start)
        if gradients[local_index] > 0:
            candidates.append(valid_radii[local_index])
    if not candidates:
        return old_radius, 0.0, 0.0
    new_radius = float(np.median(candidates))
    radius_std = float(np.std(candidates))
    valid_fraction = len(candidates) / 72.0
    confidence = max(0.0, min(1.0, valid_fraction * (1.0 - radius_std / max(new_radius, 1.0))))
    return new_radius, radius_std, confidence


def refine_radius_results(
    rgb: np.ndarray,
    rows: list[dict[str, float | int | str]],
    args: argparse.Namespace,
) -> list[dict[str, float | int | str]]:
    gray = gray_from_rgb(rgb).astype(float)
    refined_rows: list[dict[str, float | int | str]] = []
    for row in rows:
        old_radius_px = float(row["Diameter_px"]) / 2.0
        new_radius_px, radius_std, confidence = refine_radius_for_row(gray, row)
        unit_per_px = float(row.get("unit_per_px") or 1.0)
        old_diameter_um = float(row.get("Diameter_um") or old_radius_px * 2.0 * unit_per_px)
        new_diameter_um = 2.0 * new_radius_px * unit_per_px
        use_blob_fallback = (
            radius_std > args.radius_extreme_std
            and abs(new_diameter_um - old_diameter_um) > args.radius_extreme_diff_um
        )
        final_diameter_um = old_diameter_um if use_blob_fallback else new_diameter_um
        radius_status = "fallback_extreme" if use_blob_fallback else "refined"
        refined_rows.append(
            {
                "label": row["Label"],
                "old_radius_px": old_radius_px,
                "new_radius_px": new_radius_px,
                "old_diameter_um": old_diameter_um,
                "new_diameter_um": new_diameter_um,
                "Blob_Diameter_um": old_diameter_um,
                "Refined_Diameter_um": new_diameter_um,
                "Final_Diameter_um": final_diameter_um,
                "radius_std": radius_std,
                "confidence": confidence,
                "Radius_Status": radius_status,
            }
        )
    return refined_rows


def radius_refinement_sheet_rows(rows: list[dict[str, float | int | str]]) -> list[dict[str, float | int | str]]:
    return [
        {
            "label": row["label"],
            "Blob_Diameter_um": row["Blob_Diameter_um"],
            "Refined_Diameter_um": row["Refined_Diameter_um"],
            "Final_Diameter_um": row["Final_Diameter_um"],
            "radius_std": row["radius_std"],
            "confidence": row["confidence"],
            "Radius_Status": row["Radius_Status"],
        }
        for row in rows
    ]



def write_excel(
    path: Path,
    detail_rows: list[dict[str, float | int | str]],
    summary_rows: list[dict[str, float | int | str]],
    radius_refinement_rows: list[dict[str, float | int | str]] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    def add_sheet(name: str, rows: list[dict[str, float | int | str]]) -> None:
        sheet = workbook.create_sheet(name)
        if not rows:
            sheet.append(["No data"])
            return
        headers = fieldnames(rows)
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        for column_index, header in enumerate(headers, start=1):
            values = [str(header)]
            values.extend(str(row.get(header, "")) for row in rows[:200])
            width = min(max(len(value) for value in values) + 2, 32)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    add_sheet("Summary", summary_rows)
    add_sheet("Detail", detail_rows)
    if radius_refinement_rows is not None:
        add_sheet("Radius_Refinement", radius_refinement_sheet_rows(radius_refinement_rows))
    workbook.save(path)


def summarize(detail_rows: list[dict[str, float | int | str]]) -> list[dict[str, float | int | str]]:
    by_image: dict[str, list[dict[str, float | int | str]]] = {}
    for row in detail_rows:
        by_image.setdefault(str(row["Image"]), []).append(row)
    summary: list[dict[str, float | int | str]] = []
    for image, rows in by_image.items():
        ferets = np.array([float(row["Feret"]) for row in rows], dtype=float)
        summary.append(
            {
                "Image": image,
                "Count": len(rows),
                "Unit": rows[0].get("unit", "px") if rows else "px",
                "Feret_mean": float(ferets.mean()) if len(ferets) else 0.0,
                "Feret_median": float(np.median(ferets)) if len(ferets) else 0.0,
                "Feret_min": float(ferets.min()) if len(ferets) else 0.0,
                "Feret_max": float(ferets.max()) if len(ferets) else 0.0,
            }
        )
    return summary


def run(args: argparse.Namespace) -> int:
    input_path = Path(args.input)
    output_dir = Path(args.output)
    images = collect_images(input_path)
    if not images:
        print(f"No images found in {input_path}", file=sys.stderr)
        return 2

    all_rows: list[dict[str, float | int | str]] = []
    all_candidate_rows: list[dict[str, float | int | str]] = []
    all_small_candidate_rows: list[dict[str, float | int | str]] = []
    all_small_validated_rows: list[dict[str, float | int | str]] = []
    all_radius_refinement_rows: list[dict[str, float | int | str]] = []
    for image_path in images:
        rgb = load_rgb(image_path)
        rows, crop, scale_info, mask, candidate_rows = analyze_image(rgb, args)
        for row in rows:
            row["Image"] = image_path.name
            row["scale_source"] = scale_info.source
            row["scale_bar_px"] = scale_info.scale_bar_px or ""
            row["unit_per_px"] = scale_info.unit_per_px or ""
        for candidate_row in candidate_rows:
            candidate_row["candidate_id"] = len(all_candidate_rows) + 1
            all_candidate_rows.append(candidate_row)
        if args.small_blob_min_diameter_px is not None:
            small_candidate_rows = detect_small_blob_candidates(rgb, crop, scale_info, args)
            for small_candidate_row in small_candidate_rows:
                small_candidate_row["candidate_id"] = len(all_small_candidate_rows) + 1
                all_small_candidate_rows.append(small_candidate_row)
            validated_rows, accepted_rows = validate_small_blob_candidates(
                rgb,
                crop,
                scale_info,
                rows,
                small_candidate_rows,
                args,
            )
            for validated_row in validated_rows:
                validated_row["candidate_id"] = len(all_small_validated_rows) + 1
                all_small_validated_rows.append(validated_row)
            if args.merge_small_blobs:
                for accepted_row in accepted_rows:
                    accepted_row["Image"] = image_path.name
                    accepted_row["scale_source"] = scale_info.source
                    accepted_row["scale_bar_px"] = scale_info.scale_bar_px or ""
                    accepted_row["unit_per_px"] = scale_info.unit_per_px or ""
                    rows.append(accepted_row)
                    cv2.circle(
                        mask,
                        (
                            int(round(float(accepted_row["CenterX_px"]) - crop.x)),
                            int(round(float(accepted_row["CenterY_px"]) - crop.y)),
                        ),
                        int(round(float(accepted_row["Diameter_px"]) / 2.0)),
                        True,
                        -1,
                    )
        all_rows.extend(rows)
        if args.refine_radius:
            image_refinement_rows = refine_radius_results(rgb, rows, args)
            for row, refinement_row in zip(rows, image_refinement_rows):
                for key in (
                    "Blob_Diameter_um",
                    "Refined_Diameter_um",
                    "Final_Diameter_um",
                    "radius_std",
                    "confidence",
                    "Radius_Status",
                ):
                    row[key] = refinement_row[key]
            for refinement_row in image_refinement_rows:
                refinement_row["Image"] = image_path.name
                refinement_row["Image_Label"] = refinement_row["label"]
                refinement_row["label"] = len(all_radius_refinement_rows) + 1
                all_radius_refinement_rows.append(refinement_row)

        stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", image_path.stem)
        annotate_image(rgb, rows, crop, scale_info, output_dir / "annotated" / f"{stem}_pores.png")
        write_mask(mask, output_dir / "masks" / f"{stem}_mask.png")
        print(f"{image_path.name}: {len(rows)} pores")

    summary_rows = summarize(all_rows)
    write_csv(output_dir / "pores_detail.csv", all_rows)
    write_csv(output_dir / "pores_summary.csv", summary_rows)
    write_csv(output_dir / "candidate_results.csv", all_candidate_rows)
    if args.small_blob_min_diameter_px is not None:
        write_csv(output_dir / "candidate_results_small.csv", all_small_candidate_rows)
        write_csv(output_dir / "candidate_results_small_validated.csv", all_small_validated_rows)
    if args.refine_radius:
        write_csv(output_dir / "radius_refinement_results.csv", all_radius_refinement_rows)
    write_excel(
        output_dir / "pores_results.xlsx",
        all_rows,
        summary_rows,
        all_radius_refinement_rows if args.refine_radius else None,
    )
    print(f"Wrote results to {output_dir}")
    return 0


def make_self_test_image(path: Path) -> None:
    im = Image.new("RGB", (420, 300), (180, 170, 135))
    draw = ImageDraw.Draw(im)
    draw.rectangle([0, 250, 419, 299], fill=(245, 245, 245))
    for box in ([80, 70, 130, 120], [180, 90, 245, 155], [275, 155, 330, 210]):
        draw.ellipse(box, fill=(45, 35, 30))
    draw.ellipse([-20, 80, 35, 135], fill=(45, 35, 30))
    draw.ellipse([310, 30, 385, 65], fill=(45, 35, 30))
    draw.rectangle([285, 268, 365, 273], fill=(0, 0, 0))
    draw.text((300, 276), "2 um", fill=(0, 0, 0))
    path.parent.mkdir(parents=True, exist_ok=True)
    im.save(path)


def self_test(args: argparse.Namespace) -> int:
    tmp = Path(args.output) / "_self_test" / "synthetic_afm.png"
    make_self_test_image(tmp)
    args.input = str(tmp)
    args.scale_bar_value = 2.0
    args.scale_bar_unit = "um"
    args.min_diameter_px = 12.0
    args.max_diameter_px = None
    code = run(args)
    detail = Path(args.output) / "pores_detail.csv"
    if not detail.exists():
        print("Self-test failed: no detail CSV", file=sys.stderr)
        return 1
    with detail.open("r", encoding="utf-8") as handle:
        count = max(sum(1 for _ in handle) - 1, 0)
    if count != 3:
        print(f"Self-test failed: expected 3 complete circles, got {count}", file=sys.stderr)
        return 1
    print("Self-test passed")
    return code


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Measure complete circular pores in AFM images.")
    parser.add_argument("input", nargs="?", help="Image file or directory")
    parser.add_argument("--output", default="afm_results", help="Output directory")
    parser.add_argument("--crop", type=parse_box, help="Manual image crop as x,y,w,h")
    parser.add_argument("--polarity", choices=["auto", "dark", "bright"], default="dark")
    parser.add_argument("--method", choices=["auto", "region", "blob"], default="auto", help="Pore detection method")
    parser.add_argument("--blob-threshold", type=float, default=0.25, help="LoG blob detector threshold")
    parser.add_argument("--blob-overlap", type=float, default=0.4, help="Allowed overlap between blob detections")
    parser.add_argument("--scale-bar-value", type=float, help="Real scale-bar value, e.g. 4")
    parser.add_argument("--scale-bar-unit", help="Scale-bar unit, e.g. um or nm")
    parser.add_argument("--scale-bar-roi", type=parse_box, help="Manual scale-bar ROI as x,y,w,h")
    parser.add_argument("--unit-per-px", type=float, help="Direct physical-unit length per pixel")
    parser.add_argument("--unit", help="Unit used with --unit-per-px, e.g. um or nm")
    parser.add_argument("--min-diameter-px", type=float, default=20.0)
    parser.add_argument("--max-diameter-px", type=float)
    parser.add_argument("--small-blob-min-diameter-px", type=float, help="Experimental small-blob candidate minimum diameter in pixels")
    parser.add_argument("--small-blob-max-diameter-px", type=float, default=20.0, help="Experimental small-blob candidate maximum diameter in pixels")
    parser.add_argument("--merge-small-blobs", action="store_true", help="Merge accepted experimental small-blob candidates into final results")
    parser.add_argument("--small-blob-min-score", type=float, default=0.95)
    parser.add_argument("--small-blob-min-contrast", type=float, default=70.0)
    parser.add_argument("--small-blob-min-circularity-score", type=float, default=0.88)
    parser.add_argument("--refine-radius", action="store_true", help="Write experimental radial-gradient radius refinement results without changing final measurements")
    parser.add_argument("--radius-extreme-std", type=float, default=5.0)
    parser.add_argument("--radius-extreme-diff-um", type=float, default=1.0)
    parser.add_argument("--min-area-px", type=int, default=20)
    parser.add_argument("--edge-margin-px", type=int, default=18)
    parser.add_argument("--min-circularity", type=float, default=0.75)
    parser.add_argument("--min-round", type=float, default=0.80)
    parser.add_argument("--max-ar", type=float, default=1.25)
    parser.add_argument("--min-solidity", type=float, default=0.90)
    parser.add_argument("--self-test", action="store_true", help="Run a synthetic-image self-test")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.self_test:
        return self_test(args)
    if not args.input:
        parser.error("input is required unless --self-test is used")
    return run(args)


if __name__ == "__main__":
    raise SystemExit(main())

